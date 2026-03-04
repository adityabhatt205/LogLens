from __future__ import annotations

import gzip
import json
from collections.abc import AsyncIterator, Iterator
from pathlib import Path

from ..models import Event
from ..parsers.detector import FormatDetector
from ..parsers.registry import get_parser
from .base import SourceAdapter


class FileAdapter(SourceAdapter):
    """Reads log events from a single file (plain or gzip-compressed)."""

    def __init__(self, path: Path) -> None:
        self.path = path

    async def events(self) -> AsyncIterator[Event]:
        lines = self._read_lines()
        sample: list[str] = []
        buf: list[str] = []
        for line in lines:
            buf.append(line)
            if len(sample) < 5 and line.strip():
                sample.append(line)
            if len(buf) >= 512:
                break

        fmt = FormatDetector().detect(sample)
        parser = get_parser(fmt, source=str(self.path))

        for line in buf:
            event = parser.parse(line)
            if event is not None:
                yield event

        for line in self._read_lines(skip=len(buf)):
            event = parser.parse(line)
            if event is not None:
                yield event

    def _read_lines(self, skip: int = 0):
        if self.path.suffix.lower() == ".xlsx":
            yield from self._read_xlsx_lines(skip=skip)
            return
        open_fn = gzip.open if self.path.suffix == ".gz" else open
        with open_fn(self.path, "rt", encoding="utf-8", errors="replace") as f:
            for i, line in enumerate(f):
                if i < skip:
                    continue
                yield line

    def _read_xlsx_lines(self, skip: int = 0):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl not installed. Run: pip install loglens[xlsx]") from exc

        wb = load_workbook(self.path, read_only=True, data_only=True)
        try:
            for i, line in enumerate(self._xlsx_lines(wb)):
                if i >= skip:
                    yield line
        finally:
            wb.close()

    def _xlsx_lines(self, wb) -> Iterator[str]:
        """Yield one text line per data row.

        When the first sheet starts with a header row of column names, each
        data row is emitted as a JSON object keyed by those names, so the JSON
        Lines parser can pull out timestamp/level/message and populate
        parsed_fields. Sheets without a usable header fall back to tab-joined
        plain text (one cell per column).
        """
        sheets = wb.worksheets
        if not sheets:
            return

        first, *rest = sheets
        rows = first.iter_rows(values_only=True)
        first_row = next(rows, None)
        header = self._xlsx_header(first_row)

        if header is not None:
            for row in rows:
                obj = {key: ("" if cell is None else cell) for key, cell in zip(header, row)}
                yield json.dumps(obj, default=str)
        else:
            if first_row is not None:
                yield _xlsx_row_to_text(first_row)
            for row in rows:
                yield _xlsx_row_to_text(row)

        # Extra sheets are uncommon for log exports — always plain text.
        for sheet in rest:
            for row in sheet.iter_rows(values_only=True):
                yield _xlsx_row_to_text(row)

    @staticmethod
    def _xlsx_header(row) -> list[str] | None:
        """Return column names if ``row`` looks like a header, else None.

        A header is a row of at least two cells where every cell is a
        non-empty string (column titles are never numbers or blanks).
        """
        if not row or len(row) < 2:
            return None
        keys: list[str] = []
        for cell in row:
            if not isinstance(cell, str) or not cell.strip():
                return None
            keys.append(cell.strip())
        return keys


def _xlsx_row_to_text(row) -> str:
    return "\t".join("" if cell is None else str(cell) for cell in row)
